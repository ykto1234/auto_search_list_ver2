from openpyxl import Workbook, load_workbook
import os
import settings
import re
import pandas as pd
import search_list


# modeは「1：ギフトショップ、2：NAVITIME」
def out_to_excel(data_list, title_name, input_data, site_name, mode):

    # リストから店名が重複する行を削除する
    data_set = set(data_list)

    backup_name = ""

    os.makedirs('./output', exist_ok=True)
    if mode == 1:
        file_path = './output/ギフトショップ検索結果.xlsx'
    elif mode == 2:
        file_path = './output/NAVITIME検索結果.xlsx'
    elif mode == 3:
        file_path = './output/MAPION検索結果.xlsx'
    elif mode == 4:
        file_path = './output/街の情報屋さん検索結果.xlsx'
    elif mode == 5:
        file_path = './output/おでかけタウン情報検索結果.xlsx'
    elif mode == 6:
        file_path = './output/goo地図検索結果.xlsx'
    elif mode == 7:
        file_path = './output/Googleマップ検索結果.xlsx'
    else:
        return

    if os.path.exists(file_path):
        # Excelファイルが存在する場合
        wb = load_workbook(file_path)
        if title_name in wb.sheetnames:
            # 同じシート名が存在する場合
            backup_name = title_name + "_bak"
            wb[title_name].title = backup_name

        wb.create_sheet(title_name)
        target_ws = wb[title_name]

    else:
        # Excelファイルが存在しない場合、Excelワークブックの生成
        wb = Workbook()
        ws = wb.active
        ws.title = title_name
        target_ws = ws

    # ヘッダーを書き込む
    target_ws.append(['店名', '業種名', '郵便番号', '住所', '電話番号', 'URL（ウェブサイト）', '取得元サイト名'])

    # データを書き込む
    for row in data_set:
        output_flg = True

        if mode in (1, 2, 3, 4, 6):
            # 店名に除外キーワードが含まれる場合、除外する
            for exclusion_word in input_data.exclusion_genre.split(","):
                if exclusion_word and exclusion_word in row.storename:
                    output_flg = False
        elif mode == 5:
            # 店名からキーワード絞り込み
            for target_industry in input_data.target_industry_list:
                if target_industry in row.storename :
                    # 店名に絞り込み業種のキーワードが１つでも含まれる行の場合、出力する
                    output_flg = True
                    break
                else:
                    output_flg = False
        else:
            # 業種チェック
            for target_industry in input_data.target_industry_list:
                if target_industry in row.industry:
                    # 絞り込み業種に１つでも含まれる行の場合、出力する
                    output_flg = True
                    break
                else:
                    output_flg = False

        if not output_flg:
            continue

        # 郵便番号を取得する
        if mode in (1, 2, 3, 6):
            row.postal_code = "-"
            match_obj = re.search(r"\d", row.address)
            if match_obj:
                target_address = row.address[:match_obj.start()]
                postal_code = settings.read_postal_code("", "", "", target_address.replace(' ', ''))
                if postal_code and postal_code != "-":
                    row.postal_code = postal_code[:3] + "-" + postal_code[3:]

        target_ws.append([row.storename, row.industry, row.postal_code,
                  row.address, row.tel_number, row.web_url, site_name])

    if backup_name:
        del wb[backup_name]

    wb.save(file_path)


def merge_excel(excel_item_list: list):

    # データを結合する
    df_merge = pd.DataFrame(data=None, index=None, columns=None, dtype=None, copy=False)

    # 各Excelデータを読み込む
    for excel_item in excel_item_list:
        df_tmp = pd.read_excel(excel_item.out_file_path, dtype=str, sheet_name=excel_item.sheet_name, header=0, usecols="A:G")
        df_merge = df_merge.append(df_tmp)

    # データフレームから全て空白の値の行を削除する
    df_merge_formatted = df_merge.dropna(how='all', axis=0)

    # 電話番号列が空のものを取り出す
    df_telnum_null = df_merge_formatted[df_merge_formatted['電話番号'] == '-']

    # 電話番号列が空でない行に対して、電話番号の重複を削除して取り出す
    df_telnum_notnull = df_merge_formatted[df_merge_formatted['電話番号'] != '-'].drop_duplicates(['電話番号'])

    # 空行と空でない行を結合
    df_merge_formatted2 = df_telnum_notnull.append(df_telnum_null)

    # NaNのセルの値を指定
    df_ret = df_merge_formatted2.fillna('-')

    os.makedirs('./output', exist_ok=True)
    file_path = './output/検索結果_結合.xlsx'

    backup_name = ""

    target_sheet_name = excel_item_list[0].sheet_name

    if os.path.exists(file_path):
        # Excelファイルが存在する場合、既存ファイルに新規シートを書き込み
        wb = load_workbook(file_path)
        if target_sheet_name in wb.sheetnames:
            # 同じシート名が存在する場合
            backup_name = target_sheet_name + "_bak"
            wb[target_sheet_name].title = backup_name
            wb.save(file_path)
            wb.close()

        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
            df_ret.to_excel(writer, sheet_name=target_sheet_name, index=False, header=True)

        if backup_name:
            wb = load_workbook(file_path)
            del wb[backup_name]
            wb.save(file_path)
            wb.close()

    else:
        # Excelファイルが存在しない場合、新規ファイルを作成する
        df_ret.to_excel(file_path, sheet_name=target_sheet_name, index=False, header=True)

    return
